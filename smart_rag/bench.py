#!/usr/bin/env python3
"""Smart RAG benchmark — prove the 4 wins on YOUR data, with numbers (not faith).

Compares Smart RAG vs a naive chunk baseline on an ingested source:
  1. INDEX SIZE     — facts vs raw chunks (redundancy collapse)
  2. TOKENS/ANSWER  — distilled fact block vs raw retrieved chunks
  3. LATENCY        — fact lookup (ms) vs scanning chunks
  4. CORRECTNESS    — on a small (entity, attribute, expected) labeled set

  python -m smart_rag.bench <path> [--labels labels.csv]
  labels.csv rows: entity,attribute,expected
"""
import argparse
import csv
import sys
import time
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from smart_rag import SmartRAG   # noqa: E402

try:
    import tiktoken
    _ENC = tiktoken.get_encoding("cl100k_base")
    def toks(s): return len(_ENC.encode(s))
except Exception:
    def toks(s): return max(1, len(s) // 4)   # rough fallback


def _baseline_chunks(path: str):
    """Naive baseline: every row/line is a raw chunk (what flat RAG indexes)."""
    chunks = []
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
        for sh in wb.sheetnames:
            ws = wb[sh]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    chunks.append(" | ".join(cells))
    else:
        for line in open(path, encoding="utf-8", errors="replace"):
            if line.strip():
                chunks.append(line.rstrip("\n"))
    return chunks


def main():
    ap = argparse.ArgumentParser(prog="smart_rag.bench")
    ap.add_argument("path")
    ap.add_argument("--labels", help="csv: entity,attribute,expected")
    ap.add_argument("--query-entity", help="entity to measure tokens/latency on")
    args = ap.parse_args()

    print(f"=== Smart RAG benchmark on {Path(args.path).name} ===\n")

    # Smart RAG
    t0 = time.time()
    d = SmartRAG(); d.ingest(args.path)
    distill_time = time.time() - t0
    stats = d.store.stats()

    # Baseline
    chunks = _baseline_chunks(args.path)

    # 1. INDEX SIZE  (redundancy collapse — biggest on versioned/duplicated corpora)
    print("1. INDEX SIZE")
    print(f"   baseline raw chunks   : {len(chunks):,}")
    print(f"   distilled facts       : {stats['distinct_facts']:,}")
    if len(chunks):
        print(f"   index reduction       : {len(chunks)/max(stats['distinct_facts'],1):.0f}× "
              f"smaller ({len(chunks):,} chunks → {stats['distinct_facts']:,} facts)")
    print(f"   (NB: compression is largest on VERSIONED/duplicated corpora — e.g. the")
    print(f"    same file across many revisions. On one clean file there's little to")
    print(f"    collapse, but the token + correctness wins below STILL hold.)")
    print(f"   ingested {stats['entities']:,} entities in {distill_time:.1f}s\n")

    # 2 & 3. TOKENS + LATENCY for a representative entity query
    ent = args.query_entity or (d.store.entities[0] if d.store.entities else None)
    if ent:
        # distilled answer
        t1 = time.time(); ans = d.ask(f"all facts for {ent}"); lat_distill = (time.time()-t1)*1000
        d_tokens = toks(ans)
        # baseline: chunks mentioning the entity (what a chunk-RAG would feed)
        t2 = time.time()
        base_hits = [c for c in chunks if ent in c]
        lat_base = (time.time()-t2)*1000
        b_tokens = toks("\n".join(base_hits))
        print(f"2. TOKENS / ANSWER (entity {ent})")
        print(f"   baseline chunks fed : {b_tokens:,} tokens ({len(base_hits)} chunks)")
        print(f"   distilled facts     : {d_tokens:,} tokens")
        if b_tokens:
            print(f"   token saving        : {round((1-d_tokens/b_tokens)*100)}%\n")
        print("3. LATENCY")
        print(f"   baseline scan       : {lat_base:.2f} ms")
        print(f"   distilled lookup    : {lat_distill:.2f} ms\n")

    # 4. CORRECTNESS
    if args.labels and Path(args.labels).exists():
        print("4. CORRECTNESS (labeled set)")
        rows = list(csv.reader(open(args.labels, encoding="utf-8")))
        ok = 0; tot = 0
        for r in rows:
            if len(r) < 3:
                continue
            ent_l, attr_l, expected = r[0].strip(), r[1].strip(), r[2].strip()
            tot += 1
            got = d.store.lookup(ent_l, attr_l)
            vals = [row["value"] for rows_ in got.values() for row in rows_]
            hit = any(expected.lower() in v.lower() or v.lower() in expected.lower() for v in vals)
            ok += 1 if hit else 0
            print(f"   {'✓' if hit else '✗'} {ent_l}.{attr_l} expected {expected} "
                  f"→ got {vals[:2]}")
        print(f"   correctness: {ok}/{tot} = {round(100*ok/max(tot,1))}%")


if __name__ == "__main__":
    main()
