#!/usr/bin/env python3
"""Tabular adapter — Excel (.xlsx/.xlsm), CSV, JSON → Facts.

This is the proven core (validated on real product spec tables): it preserves
structure where flat chunking destroys it — auto-detecting a possibly multi-row
header, filling merged cells, finding the entity column, and decoding domain-coded
values (UFS256 → "256 GB UFS"). On the real file it recovered 97 correct facts for
a part incl. the PCB number the old RAG got WRONG.
"""
from __future__ import annotations

import csv
import json
import os
import re
from collections import defaultdict
from typing import Iterable, List, Optional

from smart_rag.adapters.base import Adapter
from smart_rag.core.fact import Fact
from smart_rag.core.value_decode import decode_value
from smart_rag.core.entity import is_entity, is_strong_entity, entity_key, entity_column_of

_DATE_RE = re.compile(r'\d{4}-\d{2}-\d{2}')
_VER_RE = re.compile(r'_(\d+(?:\.\d+)?)\.(?:xls[xm]?|csv|json)$', re.I)


def _version_from_name(name: str) -> str:
    m = _VER_RE.search(name)
    return m.group(1) if m else ""


def _maybe_date(v) -> str:
    s = str(v)
    return s[:10] if _DATE_RE.match(s) else ""


class TabularAdapter(Adapter):
    suffixes = (".xlsx", ".xlsm", ".xls", ".csv", ".json", ".parquet")
    name = "tabular"

    def extract(self, path: str) -> Iterable[Fact]:
        low = path.lower()
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            yield from self._extract_excel(path)
        elif low.endswith(".csv"):
            yield from self._extract_csv(path)
        elif low.endswith(".json"):
            yield from self._extract_json(path)
        elif low.endswith(".parquet"):
            yield from self._extract_parquet(path)

    # ── Parquet (analytics/data-lake standard) ───────────────────────────────
    def _extract_parquet(self, path: str) -> Iterable[Fact]:
        """Read .parquet via pyarrow → rows as entity-attribute facts. The first
        id-like column is the entity; other columns are attributes."""
        import os
        try:
            import pyarrow.parquet as pq
        except Exception:
            return   # pyarrow not installed → silently skip (graceful)
        try:
            tbl = pq.read_table(path)
        except Exception:
            return
        src = os.path.basename(path)
        cols = tbl.column_names
        if not cols:
            return
        from smart_rag.core.entity import is_strong_entity, entity_key
        rows = tbl.to_pylist()
        # pick an entity column: first whose values look like ids, else first col
        ent_col = next((c for c in cols
                        if any(is_strong_entity(r.get(c)) for r in rows[:50])), cols[0])
        for r in rows:
            raw = r.get(ent_col)
            if raw is None or not str(raw).strip():
                continue
            ent = entity_key(str(raw))
            for c in cols:
                if c == ent_col:
                    continue
                v = r.get(c)
                if v is not None and str(v).strip():
                    yield Fact(entity=ent, attribute=c, value=str(v), source=src)

    # ── Excel (structure-preserving) ─────────────────────────────────────────
    def _extract_excel(self, path: str) -> Iterable[Fact]:
        import openpyxl
        import warnings
        src_name = os.path.basename(path)
        ver = _version_from_name(src_name)
        # openpyxl drops unsupported formatting/validation extensions while
        # preserving the worksheet values that this adapter extracts.
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Conditional Formatting extension is not supported.*",
                category=UserWarning,
                module=r"openpyxl\.worksheet\._reader",
            )
            warnings.filterwarnings(
                "ignore",
                message="Data Validation extension is not supported.*",
                category=UserWarning,
                module=r"openpyxl\.worksheet\._reader",
            )
            wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row < 2 or ws.max_column < 2:
                continue
            self._fill_merged(ws)
            ncols = min(ws.max_column, 250)
            hdr_row = self._detect_header_row(ws, ncols=min(ncols, 80))
            headers = {c: str(ws.cell(row=hdr_row, column=c).value or "").replace("\n", " ").strip()
                       for c in range(1, ncols + 1)}
            ent_col = self._entity_column(ws, hdr_row, ncols)
            if ent_col is None:
                continue
            src = f"{src_name}::{sheet}"
            for r in range(hdr_row + 1, ws.max_row + 1):
                raw_ent = str(ws.cell(row=r, column=ent_col).value or "").strip()
                if not is_strong_entity(raw_ent):
                    continue
                ent = entity_key(raw_ent)
                for c, attr in headers.items():
                    if not attr or c == ent_col:
                        continue
                    val = ws.cell(row=r, column=c).value
                    if val is None or val == "":   # keep 0 — real value (no SXM, bit=0)
                        continue
                    yield Fact(entity=ent, attribute=attr, value=decode_value(str(val)),
                               source=src, version=ver, date=_maybe_date(val))

    @staticmethod
    def _fill_merged(ws) -> None:
        for rng in list(ws.merged_cells.ranges):
            tl = ws.cell(row=rng.min_row, column=rng.min_col).value
            if tl is None:
                continue
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    try:
                        ws.cell(row=row, column=col).value = tl
                    except Exception:
                        pass

    @staticmethod
    def _detect_header_row(ws, scan_rows: int = 12, ncols: int = 60) -> int:
        best_row, best_score = 1, -1
        for r in range(1, scan_rows + 1):
            cells = [str(ws.cell(row=r, column=c).value or "").strip()
                     for c in range(1, ncols + 1)]
            nonblank = [c for c in cells if c]
            if not nonblank or any(is_entity(c) for c in cells):
                continue
            labelish = sum(1 for c in nonblank
                           if 1 < len(c) <= 40 and re.search(r'[A-Za-z]', c))
            score = labelish + len(set(nonblank))
            if score > best_score:
                best_score, best_row = score, r
        return best_row

    @staticmethod
    def _entity_column(ws, hdr_row: int, ncols: int, sample: int = 80) -> Optional[int]:
        # Prefer the column richest in STRONG entities (real domain ids), not the
        # generic 4-digit-code columns that pollute wide engineering sheets.
        strong = defaultdict(int)
        for r in range(hdr_row + 1, min(hdr_row + 1 + sample, ws.max_row + 1)):
            for c in range(1, ncols + 1):
                if is_strong_entity(ws.cell(row=r, column=c).value):
                    strong[c] += 1
        if strong:
            return max(strong, key=strong.get)
        return None

    # ── CSV ──────────────────────────────────────────────────────────────────
    def _extract_csv(self, path: str) -> Iterable[Fact]:
        src = os.path.basename(path)
        with open(path, encoding="utf-8", errors="replace", newline="") as fh:
            rows = list(csv.reader(fh))
        if len(rows) < 2:
            return
        header = [h.strip() for h in rows[0]]
        ent_col = entity_column_of(rows[1:])
        for row in rows[1:]:
            if ent_col is None or ent_col >= len(row):
                continue
            raw_ent = row[ent_col].strip()
            if not raw_ent:
                continue
            ent = entity_key(raw_ent) if is_entity(raw_ent) else raw_ent
            for i, attr in enumerate(header):
                if i == ent_col or not attr or i >= len(row) or not row[i].strip():
                    continue
                yield Fact(entity=ent, attribute=attr, value=decode_value(row[i].strip()),
                           source=src, date=_maybe_date(row[i]))

    # ── JSON (list of objects, or dict of objects) ───────────────────────────
    def _extract_json(self, path: str) -> Iterable[Fact]:
        src = os.path.basename(path)
        try:
            data = json.loads(open(path, encoding="utf-8", errors="replace").read())
        except Exception:
            return
        records = data if isinstance(data, list) else (
            list(data.values()) if isinstance(data, dict) else [])
        for rec in records:
            if not isinstance(rec, dict):
                continue
            # entity = first id-looking field, else first string field
            ent = None
            for k, v in rec.items():
                if isinstance(v, (str, int)) and ("id" in k.lower() or "name" in k.lower()
                                                  or "number" in k.lower() or is_entity(v)):
                    ent = str(v); break
            if ent is None:
                ent = str(next((v for v in rec.values() if isinstance(v, (str, int))), ""))
            if not ent:
                continue
            ent = entity_key(ent) if is_entity(ent) else ent
            for k, v in rec.items():
                if isinstance(v, (dict, list)) or v in (None, ""):
                    continue
                yield Fact(entity=ent, attribute=str(k), value=decode_value(str(v)),
                           source=src, date=_maybe_date(v))
